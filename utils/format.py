###################################################格式化报文，去除多余空行########################################################
#
#python .\utils\format.py --input_file "D:\Downloads\1.xlsx" --row 2 --column 16 --output_file "D:\Downloads\2.xlsx" --output_column 23
#
#################################################################################################################################

import time
import argparse
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class Formatter:
    """报文格式化工具"""

    def format(self, text):
        """格式化文本，去除多余的空行"""
        #去除多余换行
        lines = text.splitlines()  # 按行分割，自动去除每行末尾的换行符
        result = []

        in_consecutive_empty = False  # 标记当前是否在连续空值段中

        empty_values = (None, '')

        for item in lines:
            if item in empty_values:
                if not in_consecutive_empty:
                    # 第一个空值，跳过（即删除）
                    in_consecutive_empty = True
                else:
                    # 连续的后续空值，保留
                    result.append(item)
            else:
                # 非空值，正常添加，并重置标记
                result.append(item)
                in_consecutive_empty = False

        formatted_text = '\n'.join(result)
        return formatted_text

    def format_excel(self, input_file, row, column, output_file, output_column, log_func=None):
        """格式化Excel文件中的报文"""
        log = log_func if log_func else print
        
        try:
            wb = load_workbook(filename=input_file)
            ws = wb['Sheet1']  # 输入表

            # 输出文件，如果不存在复制源文件，如果存在打开
            try:
                wb1 = load_workbook(output_file)
            except FileNotFoundError:
                wb1 = wb

            ws1 = wb1.active

            # 定义遍历的行范围
            start_row = row  # 开始行（包含），基于1索引
            col = column   # 报文所在列
            col1 = output_column    # 输出的列

            # 获取最大行号
            max_row = ws.max_row

            for row in ws.iter_rows(min_row=start_row, max_row=max_row, min_col=col, max_col=col, values_only=True):
                if any(cell is not None for cell in row):  # 检查行是否包含非空单元格
                    tmp_row = row[0]
                    formatted_text = self.format(tmp_row)
                    ws1.cell(row=start_row, column=col1, value=formatted_text)
                start_row += 1

            ws1.column_dimensions[get_column_letter(col)].width = 100  # 设置报文所在列宽
            wb1.save(output_file)
            log("格式化完成！")
        except Exception as e:
            log(f"发生错误: {e}")


def main():
    parser = argparse.ArgumentParser(description='报文格式化  v1.8.3')
    parser.add_argument('--input_file', type=str, help='输入文件', required=True)
    parser.add_argument('--row', type=int, help='第一个报文所在行号', required=True)
    parser.add_argument('--column', type=int, help='报文所在列号', required=True)
    parser.add_argument('--output_file', type=str, help='输出文件', required=True)
    parser.add_argument('--output_column', type=int, help='输出列', required=True)

    args = parser.parse_args()
    input_file = args.input_file
    row = args.row
    column = args.column
    output_file = args.output_file
    output_column = args.output_column

    formatter = Formatter()
    print("转换中，请稍候......")
    formatter.format_excel(input_file, row, column, output_file, output_column)


if __name__ == '__main__':
    main()