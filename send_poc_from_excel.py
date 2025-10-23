import time
import argparse
import requests
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SEND_INTERVAL=0.1   #发送间隔
STOP_TO_SEND=0

class excel_sender:

    def __init__(self,input_file,row,column,output_file,output_column,dst, use_https,log_func=None, finish_callback=None):
        self.input_file=input_file
        self.row=row
        self.column=column
        self.output_file=output_file
        self.output_column=output_column
        self.dst=dst
        self.log = log_func if log_func else print  # 使用传入的 log 函数，否则默认 print
        self.finish = finish_callback if finish_callback else lambda: None  # 完成回调
        self.use_https=use_https       #----------------v20251022，增加https参数-----------------#


    def parse(self,raw_request):
        #解析一条poc
        # 解码URL编码的字符
        decoded_request = raw_request
        decoded_request = decoded_request.replace('_x000D_', '')     # v1.5 excel中的换行符_x000D_
        decoded_request = decoded_request.replace('\u200b', '')      # v1.8零宽度空格

        if '\r\n\r\n' in decoded_request:
            header_body_split = decoded_request.split('\r\n\r\n')
        else:
            header_body_split = decoded_request.split('\n\n')

        headers_str = header_body_split[0]
        # v1.7，ct支持br压缩，request默认不支持，直接替换，否则无法解析响应
        headers_str = re.sub(r'accept-encoding:\s*.*', 'accept-encoding: gzip, deflate', headers_str, flags=re.IGNORECASE)

        body_str = ""

        for i in range(len(header_body_split)):
            if i==1:
                body_str=body_str + header_body_split[i]
            if i>1:
                body_str=body_str + "\n\n" + header_body_split[i]     #增加请求体多个空行处理，v1.1

        body_str=body_str.replace('\n','\r\n')        #将换行统一替换成\r\n,v1.6

        # 解析请求头
        self.headers = {}
        self.body=""
        self.method=""
        self.path=""
        self.real_content_length = 0

        ##########v1.8.2，重构，调整匹配顺序，支持请求头冒号后面没有空格，冒号后面没有值####################
        request_line = headers_str.splitlines()[0]
        match = re.match(r'^(GET|POST|PUT|DELETE|HEAD|OPTIONS|PATCH|TRACE|CONNECT)\s+([^\s]+)\s+HTTP/([\d.]+)$',
                         request_line)
        if match:
            self.method, self.path, self.http_version = match.groups()

        for line in headers_str.splitlines()[1:]:
            line = line.strip()  # 建议加 strip
            if re.match(r'^[^:]+:$', line):  # 更精确：以 : 结尾，且前面无空格
                key, value = line.split(':', 1)
                self.headers[key] = value  # value == ""
            elif ': ' in line:
                key, value = line.split(': ', 1)
                self.headers[key] = value
            elif ':' in line:
                key, value = line.split(':', 1)
                self.headers[key] = value.strip()  # 建议 strip

        self.body = body_str

        self.real_content_length = len(body_str)

        ##### v20251022，body为空但存在content-length时，删除content-length。body不为空时，python已经自动处理
        if self.real_content_length == 0:
            # 删除所有大小写变体的 Content-Length 头
            keys_to_delete = [k for k in self.headers if k.lower() == 'content-length']
            for k in keys_to_delete:
                del self.headers[k]


        # ==================== v1.8.2：自动添加 User-Agent，避免自动添加python-request头，触发爬虫规则 ====================
        # 检查是否已存在 User-Agent（忽略大小写）
        has_user_agent = any(key.lower() == 'user-agent' for key in self.headers.keys())
        if not has_user_agent:
            # 可以自定义你喜欢的 User-Agent
            self.headers[
                'User-Agent'] = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        # ==================== 结束新增 ====================

    # import requests

    def send_1_poc(self):
        #发送一条poc
        # self.return_content="
        # print(self.method)

        # v20251022支持https功能
        if self.use_https in ['yes', 'y', 'Yes', 'Y']:
            url = f"https://{self.dst}{self.path}"
        else:
            url = f"http://{self.dst}{self.path}"
        response = requests.request(
            method=self.method,         #v1.8.3,选择报文里的请求方法
            url=url,
            headers=self.headers,
            data=self.body,
            allow_redirects=False,        #v1.8.1禁止重定向，否则重定向次数过多会发送失败
            verify=False
        )

        return response

    def read_excel(self):
        try:
            file_path = self.input_file  # 替换为你的Excel文件路径
            output_file_path = self.output_file
            sheet_name = 'Sheet1'  # 替换为你想要读取的工作表名称

            wb = load_workbook(filename=file_path)
            ws = wb[sheet_name]    #输入表

        #v1.8，输出文件，如果不存在复制源文件，如果存在打开，可以中断之后再执行不会覆盖原来的数据。
            try:
                wb1 = load_workbook(output_file_path)
            except FileNotFoundError:
                wb1 = wb

            ws1 = wb1.active

            # 定义遍历的行范围
            start_row = self.row  # 开始行（包含），基于1索引
            col=self.column   #报文所在列
            col1=self.output_column    #输出的列，

            # 获取最大行号
            max_row = ws.max_row

            for row in ws.iter_rows(min_row=start_row, max_row=max_row,  min_col=col, max_col=col,values_only=True):
                if STOP_TO_SEND==1:
                    self.log("发送已停止")
                    print("发送已停止")  # 20251011,增加暂停功能
                    # STOP_TO_SEND=0
                    break

                if any(cell is not None for cell in row):  # 检查行是否包含非空单元格
                    tmp_row=row[0]
                    self.parse(tmp_row)        #执行解析不能删，返回是为了测试

                    try:             # v1.4,异常退出并报错
                        response_all=self.send_1_poc()
                    except Exception as e:
                        self.log("第" + str(start_row) + "行发送异常，请检查后重新发送！")
                        print("第" + str(start_row) + "行发送异常，请检查后重新发送！")
                        break

                    ws1.cell(row=start_row, column=col1, value=response_all.text)
                    ws1.cell(row=start_row,column=col1+1,value=response_all.status_code)

                    pattern_ah = r"Request ID:\s*(\d{19})"                 #v1.8.1，增加空格，适应545版本
                    pattern_ct = r"event_id:\s*([a-fA-F0-9]{32})"              #v1.3,ct返回id
                    match_ah = re.search(pattern_ah, response_all.text)
                    if match_ah:
                        ws1.cell(row=start_row, column=col1+2, value=match_ah.group(1))

                    match_ct = re.search(pattern_ct, response_all.text)
                    if match_ct:
                        ws1.cell(row=start_row, column=col1 + 2, value=match_ct.group(1))

                    # wb1.save(output_file_path)  #v1.8.1 频繁保存大文件太慢

                else:         #v1.5 解决结束不能退出问题
                    # wb1.save(output_file_path)
                    break
                self.log("第"+str(start_row)+"行发送完成，响应码："+str(response_all.status_code))
                print("第"+str(start_row)+"行发送完成，响应码："+str(response_all.status_code))      #v1.4，打印行号    v1.7，打印响应码
                start_row += 1
                time.sleep(SEND_INTERVAL)    #v1.8.1 太慢，后面考虑增加输入参数
                # print(STOP_TO_SEND)

            ws1.column_dimensions[get_column_letter(col)].width=100    #v1.8设置报文所在列宽
            wb1.save(output_file_path)
        except Exception as e:
            self.log(f"发生错误: {e}")
        finally:
            self.finish()


def main():

    parser = argparse.ArgumentParser(description='POC批量发送  v1.8.3')
    parser.add_argument('--input_file', type=str, help='输入文件',required=True)
    parser.add_argument('--row', type=int, help='第一个报文所在行号',required=True)
    parser.add_argument('--column', type=int, help='报文所在列号',required=True)
    parser.add_argument('--output_file', type=str, help='输出文件', required=True)
    parser.add_argument('--output_column', type=int, help='响应内容、状态码、Request ID，输出三列', required=True)
    parser.add_argument('--dst', type=str, help='站点地址，ip:port，不支持https', required=True)
    #----------------v20251022，增加https参数-----------------#
    parser.add_argument('--use_https', type=str, help='是否使用https，Yes/No?', required=False,default='No',choices=['yes', 'no', 'y', 'n', 'Yes', 'No', 'Y', 'N'])

    args = parser.parse_args()
    input_file=args.input_file
    row=args.row
    column=args.column
    output_file=args.output_file
    output_column=args.output_column
    dst=args.dst
    https=args.use_https

    test=excel_sender(input_file,row,column,output_file,output_column,dst,https)
    print("检测中，请稍候......")
    test.read_excel()
    print("检测完成！")

if __name__ == '__main__':
    main()
