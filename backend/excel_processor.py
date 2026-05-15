import time
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from .http_client import HTTPClient
from .request_parser import RequestParser

class ExcelProcessor:
    """Excel 文件处理器"""
    
    def __init__(self, input_file, row, column, output_file, output_column, dst, use_https, 
                 log_func=None, finish_callback=None, send_interval=0.1):
        self.input_file = input_file
        self.row = row
        self.column = column
        self.output_file = output_file
        self.output_column = output_column
        self.dst = dst
        self.use_https = use_https
        self.log = log_func if log_func else print
        self.finish = finish_callback if finish_callback else lambda: None
        self.send_interval = send_interval
        self.stop_flag = False
        
        self.http_client = HTTPClient()
        self.parser = RequestParser()
    
    def stop(self):
        """停止处理"""
        self.stop_flag = True
    
    def process(self):
        """处理 Excel 文件并发送请求"""
        try:
            # 加载 Excel 文件
            wb = load_workbook(filename=self.input_file)
            ws = wb['Sheet1']
            
            # 处理输出文件
            try:
                wb1 = load_workbook(self.output_file)
            except FileNotFoundError:
                wb1 = wb
            
            ws1 = wb1.active
            start_row = self.row
            col = self.column
            col1 = self.output_column
            max_row = ws.max_row
            
            for row in ws.iter_rows(min_row=start_row, max_row=max_row, 
                                   min_col=col, max_col=col, values_only=True):
                if self.stop_flag:
                    self.log("处理已停止")
                    break
                
                if any(cell is not None for cell in row):
                    tmp_row = row[0]
                    try:
                        # 解析请求
                        request_data = self.parser.parse(tmp_row)
                        
                        # 构建 URL
                        if self.use_https in ['yes', 'y', 'Yes', 'Y']:
                            url = f"https://{self.dst}{request_data['path']}"
                        else:
                            url = f"http://{self.dst}{request_data['path']}"
                        
                        # 发送请求
                        response = self.http_client.send_request(
                            method=request_data['method'],
                            url=url,
                            headers=request_data['headers'],
                            body=request_data['body']
                        )
                        
                        # 写入响应
                        ws1.cell(row=start_row, column=col1, value=response.text)
                        ws1.cell(row=start_row, column=col1+1, value=response.status_code)
                        
                        # 提取 Request ID
                        pattern_ah = r"Request ID:\s*(\d{19})"
                        pattern_ct = r"event_id:\s*([a-fA-F0-9]{32})"
                        match_ah = re.search(pattern_ah, response.text)
                        if match_ah:
                            ws1.cell(row=start_row, column=col1+2, value=match_ah.group(1))
                        
                        match_ct = re.search(pattern_ct, response.text)
                        if match_ct:
                            ws1.cell(row=start_row, column=col1+2, value=match_ct.group(1))
                        
                        self.log(f"第{start_row}行发送完成，响应码：{response.status_code}")
                    except Exception as e:
                        self.log(f"第{start_row}行处理失败：{str(e)}")
                        ws1.cell(row=start_row, column=col1, value=f"处理失败：{str(e)}")
                        ws1.cell(row=start_row, column=col1+1, value=0)
                
                start_row += 1
                time.sleep(self.send_interval)
            
            # 保存结果
            ws1.column_dimensions[get_column_letter(col)].width = 100
            wb1.save(self.output_file)
        except Exception as e:
            self.log(f"发生错误: {e}")
        finally:
            self.finish()